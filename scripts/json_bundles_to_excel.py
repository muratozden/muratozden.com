#!/usr/bin/env python3
"""
Read data/manifest.json bundles, merge JSON parts, export one .xlsx per bundle to excel/.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / "data"
MANIFEST = DATA / "manifest.json"
OUT_DIR = REPO / "excel"

INVALID_SHEET_CHARS = re.compile(r"[\[\]\*\?\:/\\]")


def load_bundle_rows(bundle: dict) -> list[dict]:
    rows: list[dict] = []
    for fn in bundle["files"]:
        path = DATA / fn
        if not path.is_file():
            raise FileNotFoundError(path)
        doc = json.loads(path.read_text(encoding="utf-8"))
        part = doc.get("data")
        if not isinstance(part, list):
            continue
        rows.extend(part)
    return rows


def format_cell(val) -> str:
    if val is None:
        return ""
    if isinstance(val, list):
        return "; ".join(format_cell(x) if not isinstance(x, (dict, list)) else json.dumps(x, ensure_ascii=False) for x in val)
    if isinstance(val, (dict,)):
        return json.dumps(val, ensure_ascii=False)
    return str(val)


def flatten_item(item: dict) -> dict[str, str]:
    flat: dict[str, str] = {}
    for key in ("url", "canonicalUrl", "h1", "title"):
        if key in item:
            flat[key] = format_cell(item.get(key))
    attrs = item.get("attributes")
    if isinstance(attrs, dict):
        for group, props in attrs.items():
            if not isinstance(props, dict):
                continue
            for attr_key, attr_val in props.items():
                col = f"{group} / {attr_key}"
                flat[col] = format_cell(attr_val)
    return flat


def collect_columns(items: list[dict]) -> tuple[list[str], list[dict[str, str]]]:
    """Stable column order: base keys first, then sorted attribute columns."""
    base = ["url", "canonicalUrl", "h1", "title"]
    attr_cols: set[str] = set()
    flats: list[dict[str, str]] = []
    for item in items:
        f = flatten_item(item)
        flats.append(f)
        for k in f:
            if k not in base:
                attr_cols.add(k)
    ordered_attr = sorted(attr_cols, key=lambda s: s.lower())
    headers = [h for h in base if any(h in fl for fl in flats)] + ordered_attr
    return headers, flats


def safe_sheet_title(bundle_id: str, max_len: int = 31) -> str:
    """Excel worksheet name limit 31 chars; forbidden chars replaced."""
    s = INVALID_SHEET_CHARS.sub("_", bundle_id)[:max_len]
    return s or "sheet"


def write_xlsx(path: Path, sheet_title: str, headers: list[str], rows: list[dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    for r, flat in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=flat.get(h, ""))

    # Narrow reasonable default width (optional)
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        ws.column_dimensions[col].width = min(48, max(12, len(headers[c - 1]) // 2 + 2))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    bundles = manifest.get("bundles")
    if not isinstance(bundles, list):
        raise SystemExit("manifest.json: missing bundles[]")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for b in bundles:
        bid = b.get("id")
        files = b.get("files")
        if not bid or not isinstance(files, list) or not files:
            continue
        items = load_bundle_rows(b)
        if not items:
            print(f"skip empty: {bid}")
            continue
        headers, flats = collect_columns(items)
        out_path = OUT_DIR / f"{bid}.xlsx"
        write_xlsx(out_path, safe_sheet_title(bid), headers, flats)
        print(f"wrote {out_path.name} ({len(items)} rows, {len(headers)} cols)")

    print(f"done -> {OUT_DIR}")


if __name__ == "__main__":
    main()
